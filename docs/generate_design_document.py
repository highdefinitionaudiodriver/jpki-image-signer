"""
JPKI Image Signer - 設計書生成スクリプト (v0.1.0 / Phase 4 完了版)

更新履歴:
  - 初版           : Phase 1 着手時点(test_01/02 実装済み)
  - Phase 1 完了版 : test_01〜04 実装+実機検証済
  - Phase 2 完了版 : jpki/crypto/container/cli モジュール完成、
                     40件のユニットテスト + 実カードE2Eテスト Pass
  - Phase 3 完了版 : PyQt6 GUI 実装。D&D + モード自動切替 + QThread
                     非同期署名・検証 + PIN同期 + SAN OtherName 氏名抽出
  - v0.1.0 (Phase 4 完了): マルチOS対応 build.py、PyInstaller化、
                            アイコン自動生成 (PNG→ICO/ICNS)、配布ZIP化、
                            プライバシー警告ダイアログ、Antivirus対応ドキュメント
                            ★現在 (リリース可能状態)

実行:
    cd C:\\dev\\jpki-image-signer\\docs
    py -3.12 generate_design_document.py

依存:
    py -3.12 -m pip install openpyxl
"""
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl がインストールされていません。", file=sys.stderr)
    print("       py -3.12 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = SCRIPT_DIR / "design_document.xlsx"


# ===== スタイル =====
TITLE_FONT  = Font(bold=True, size=14, color="000000")
NOTE_FONT   = Font(italic=True, size=10, color="666666")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DONE_FILL   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 薄緑
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(ws, title, note, headers, rows, col_widths,
                header_row_height=30, status_col_index=None):
    """共通シート書き込み.

    status_col_index: 「実装済み」セルを薄緑で塗る列番号(1始まり)。Noneで無効。
    """
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws.cell(row=2, column=1, value=note).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.row_dimensions[2].height = 30

    header_row = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[header_row].height = header_row_height

    for r_idx, row in enumerate(rows, header_row + 1):
        is_done = False
        if status_col_index is not None and len(row) >= status_col_index:
            val = row[status_col_index - 1]
            if isinstance(val, str) and "実装済み" in val:
                is_done = True
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = CELL_ALIGN
            c.border = BORDER
            if is_done:
                c.fill = DONE_FILL

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # =============================================================
    # Sheet 1: 機能一覧表
    # =============================================================
    ws1 = wb.create_sheet("1.機能一覧表")
    headers1 = [
        "機能ID", "機能名", "概要", "対象ユーザー",
        "実装フェーズ", "実装状況", "関連ファイル", "備考",
    ]
    rows1 = [
        # ---- Phase 1 (CLI 実機検証用) ----
        ("F-001", "リーダー疎通確認",
         "PC/SC経由でICカードリーダーを認識し、JPKI AP(AID: D3 92 F0 00 26 01 00 00 00 01)へのSELECTが成功するかを確認する。PIN不要・非破壊。",
         "開発者", "Phase 1", "実装済み",
         "phase1/test_01_connect.py",
         "ATR取得+AP SELECTのみ。実機検証済(VID 058F PID 9540 / Alcor AU9540 系接触型リーダーで動作確認)。"),

        ("F-002", "利用者証明用電子証明書の読み出し",
         "EF=0x000A をSELECTし、READ BINARYで利用者証明用電子証明書(X.509 DER)を取得する。PIN不要。",
         "開発者", "Phase 1", "実装済み",
         "phase1/test_02_read_cert.py",
         "実機で 3808バイトEFから実DER 1574B(Subject CN=識別符号)を取得・openssl/certutil検証成功。"),

        ("F-003", "PIN残回数の事前確認",
         "署名用PIN EF (0x001B)をSELECTし、データ無し VERIFY を送出。SW=63CX のXから残試行回数を取得する。",
         "開発者", "Phase 1", "実装済み",
         "phase1/test_03_sign.py / phase2/jpki/session.py:JpkiSession.get_pin_remaining()",
         "Case1 4-byte / Case3 Lc=0 5-byte の2バリエーション対応。Phase 2 で JpkiSession.assert_safe_to_attempt_pin() に昇格(残回数<3で例外)。"),

        ("F-004", "JPKI署名実行",
         "対象データのSHA-256ダイジェストをDigestInfo(51バイト)として構築し、VERIFY PIN後にCOMPUTE DIGITAL SIGNATURE(80 2A 00 80)で256バイトのRSA-2048署名を取得する。",
         "開発者→エンドユーザー", "Phase 1", "実装済み",
         "phase1/test_03_sign.py / phase2/jpki/session.py:JpkiSession.sign_digest_info()",
         "PIN必須。Phase 2 でクラス化、安全装置(残回数<3で自動中止/getpass/PINメモリゼロクリア/再試行禁止)継承。"),

        ("F-004B", "署名用電子証明書の読み出し",
         "PIN認証成功後、同一セッション内で署名用電子証明書 EF=0x0001 を取得する。",
         "開発者", "Phase 1", "実装済み",
         "phase1/test_03_sign.py / phase2/jpki/session.py:JpkiSession.read_sign_certificate()",
         "実機で実DER 1749B(EF確保 3808B)取得。氏名・住所・生年月日・性別を含むため取扱注意。"),

        ("F-004C", "ダミー署名の数学的検証",
         "test_03 の出力を sign_cert.der の公開鍵で RSASSA-PKCS1-v1_5 + SHA-256 として検証し、JPKI署名の数学的妥当性を確認する。",
         "開発者", "Phase 1", "実装済み",
         "phase1/test_04_verify_dummy.py",
         "実機検証で 'OK 署名は有効' を確認。ASN.1 SEQUENCEヘッダから実DER長算出してEFパディング除去。"),

        # ---- Phase 2 (再利用モジュール) ----
        ("F-005", "PKCS#7分離署名(p7s)生成",
         "JPKIカードからの256B署名値とX.509証明書(DER)を入力に、CMS SignedData(detached)構造のp7sファイルを生成する。",
         "エンドユーザー", "Phase 2", "実装済み",
         "phase2/crypto/p7s.py:build_p7s()",
         "asn1crypto.cmsで手動構築。signedAttrs 無しの最小構成。version='v1' (PKCS#7 v1.5互換)。実機検証で2,250B出力を確認。"),

        ("F-006", "PKCS#7分離署名(p7s)検証",
         "p7sと画像本体から、署名・証明書・ハッシュ整合性を検証する。改ざん検知含む。",
         "エンドユーザー", "Phase 2", "実装済み",
         "phase2/crypto/p7s.py:verify_p7s_against_data()",
         "cryptography.hazmat の RSAPublicKey.verify を使用(内部でDigestInfo構築+PKCS1パディング+RSA復号)。22ユニットテストPass済。1MBデータの1bit反転で改ざん検知確認済。"),

        ("F-006B", "高レベル検証ラッパー + JPKI氏名抽出",
         "画像と.p7sから検証結果(valid/署名者氏名/有効期間)を辞書で返す。Phase 3 Step1(v2)で JPKI SAN OtherName(OID 1.2.392.200149.8.5.5.1) から漢字氏名を抽出する extract_signer_name() を追加実装。",
         "エンドユーザー", "Phase 2/3", "実装済み",
         "phase2/crypto/verify.py:verify_signed_image() / extract_signer_name()",
         "抽出順位: ①SAN OtherName(JPKI規格) → ②SAN DirectoryName.CN → ③Subject CN(JPKI識別符号)。実機 .jpkiimg で漢字氏名表示成功。"),

        ("F-007", ".jpkiimg コンテナ作成",
         "元画像/signature.p7s/cert.der を無圧縮ZIP(ZIP_STORED)に格納し、.jpkiimg として保存する。",
         "エンドユーザー", "Phase 2", "実装済み",
         "phase2/container/writer.py:create_jpkiimg()",
         "元画像は無加工(再エンコード禁止)。実機テストでJPEG 30,634B → コンテナ 34,957B(=image+p7s+cert+ZIPオーバーヘッド)を確認。拡張子は小文字化、無拡張子は.binにフォールバック。"),

        ("F-008", ".jpkiimg コンテナ検証",
         "コンテナ展開・必須エントリ確認・p7s検証 を行う。",
         "エンドユーザー", "Phase 2", "実装済み",
         "phase2/container/reader.py:read_jpkiimg() + phase2/crypto/verify.py",
         "18ユニットテストPass(改ざん検知/異常系含む)。ZIPでない/必須エントリ欠落でNotJpkiImgError/MissingEntryError。"),

        ("F-008B", "サンプル画像生成",
         "テスト用の汎用ダミーJPEGを生成する(個人情報を含まない)。",
         "開発者", "Phase 2", "実装済み",
         "docs/make_sample_image.py",
         "Pillow使用。SteelBlue背景に「JPKI Test Image」のテキスト + 注意書き。"),

        ("F-008C", "CLI: 署名(sign_image)",
         "コマンドラインで画像→.jpkiimgを生成する。8ステップの進捗表示+ANSIカラー+--no-colorオプション。",
         "エンドユーザー", "Phase 2", "実装済み",
         "phase2/cli/sign_image.py",
         "実カードと連動。assert_safe_to_attempt_pin()→getpass→VERIFY→sign→cert読出→build_p7s→create_jpkiimgの完全フロー。実機E2Eテスト成功。"),

        ("F-008D", "CLI: 検証(verify_image)",
         "コマンドラインで.jpkiimgを検証して結果を表示する。",
         "エンドユーザー", "Phase 2", "実装済み",
         "phase2/cli/verify_image.py",
         "成功時=緑+✅、失敗時=赤+❌。実機E2Eテストで正常系/改ざん版両方の判定を確認。"),

        # ---- Phase 3 (GUI) ----
        ("F-009", "GUI: 署名モード",
         "画像ファイルD&D → PIN入力ダイアログ → 署名実行 → .jpkiimg保存 → 結果表示。",
         "エンドユーザー", "Phase 3", "実装済み",
         "phase3/gui/sign_panel.py (SignPanel)",
         "PyQt6 採用。SignWorker(QThread)で非同期実行。QMutex+QWaitCondition で UI⇄ワーカー間の PIN同期。実機E2E成功。"),

        ("F-010", "GUI: 検証モード",
         ".jpkiimgファイルD&D → 検証 → 「有効/改ざん/不正コンテナ」を3パターンのカードUIで表示。",
         "エンドユーザー", "Phase 3", "実装済み",
         "phase3/gui/verify_panel.py (VerifyPanel)",
         "VerifyWorker(QThread)で非同期実行。verify_signed_image() の結果を緑/赤/橙カードに分岐。SAN OtherName 氏名表示対応。"),

        ("F-009A", "GUI: ウィンドウ全体D&D + 拡張子による自動モード切替",
         "QMainWindow のドロップイベントで JPEG/PNG → 署名 / .jpkiimg → 検証 を自動判別。"
         " ホバー時に背景色・点線枠が変化(視覚フィードバック)。複数同時ドロップや非対応拡張子は警告ダイアログで拒否。",
         "エンドユーザー", "Phase 3", "実装済み",
         "phase3/gui/main_window.py (MainWindow + WelcomePanel)",
         "QStackedWidget による Welcome/Sign/Verify の3画面遷移。動的プロパティ 'dragging' で QSS をホットスワップ。"),

        ("F-009B", "GUI: PinDialog (PIN入力モーダル)",
         "残回数を色分け表示(5=緑/3〜4=黄+警告/2以下=赤+警告)。QLineEdit(Password) + リアルタイム桁数バリデーション。"
         " 認証ボタンは正しい桁数になるまで disabled。閉じる時に必ず入力欄を clear。",
         "エンドユーザー", "Phase 3", "実装済み",
         "phase3/gui/pin_dialog.py (PinDialog)",
         "Windows 11 ダークモード環境でも常にライトテーマで描画されるよう QPalette を明示設定。"),

        ("F-009C", "GUI: バックグラウンドワーカー基盤",
         "SignWorker / VerifyWorker は QThread を継承し、stage_started / pin_needed / "
         "result_ready / error_occurred のシグナルで UIスレッドと通信。",
         "(内部基盤)", "Phase 3", "実装済み",
         "phase3/gui/workers.py",
         "PIN受け渡しは QMutex+QWaitCondition で双方向同期。cancel() で安全中断。Stage毎に "
         "isInterruptionRequested() を確認して途中キャンセル可能。"),

        ("F-009D", "GUI: 出力先フォルダを開く",
         "署名成功時のカードに「📂 出力先フォルダを開く」ボタンを表示。"
         " Windows では explorer /select でファイル選択した状態でエクスプローラを開く。",
         "エンドユーザー", "Phase 3", "実装済み",
         "phase3/gui/sign_panel.py:_open_output_folder()",
         "Windows以外は QDesktopServices で親フォルダを開くフォールバック。"),

        ("F-009E", "GUI: モダンフラットスタイル + ANSIライトテーマ強制",
         "Tailwind系カラーパレット(emerald/blue/amber/red)で統一されたQSS。"
         " 結果カードは success/error/warning の3バリエーション(角丸+薄色背景+太枠)。",
         "エンドユーザー", "Phase 3", "実装済み",
         "phase3/gui/styles.py (APP_STYLESHEET)",
         "システムダークモード時の見えにくさを QPalette + QDialog QSS で抑止。"),

        ("F-010A", "GUI: 検証結果の3パターン分岐表示",
         "valid=True → 緑カード(✅ 有効な署名です + 署名者・有効期間・画像情報)、"
         " valid=False → 赤カード(❌ 改ざんを検知しました)、"
         " error あり → 橙カード(⚠ 不正なコンテナです + 原因種別 + 詳細)。",
         "エンドユーザー", "Phase 3", "実装済み",
         "phase3/gui/verify_panel.py:_show_*_card()",
         "原因種別は VerifyWorker が NotJpkiImg / MissingEntry / FileNotFound 等を分類。"),

        # ---- Phase 4 (配布) ----
        ("F-011", "PyInstaller によるパッケージング",
         "PyInstaller で .exe / .app / Linux バイナリを生成。--onedir(推奨) と --onefile を選択可。"
         " --noconsole / --windowed でGUIアプリとして起動(コンソール非表示)。",
         "エンドユーザー", "Phase 4", "実装済み",
         "build.py:build_pyinstaller_args()",
         "PyQt6 / cryptography / pyscard / asn1crypto のhidden imports + DLL収集。実機E2E確認済(Win11 64bit)。"),

        ("F-011A", "マルチOS対応ビルドスクリプト",
         "platform.system() で Windows/macOS/Linux を判定し、各OS向けの PyInstaller フラグ・"
         "アイコン形式・実行ファイル形式 (.exe/.app/Linux ELF) を自動切替する。",
         "開発者", "Phase 4", "実装済み",
         "build.py (detect_os, prepare_icon, build_pyinstaller_args)",
         "macOS ICNS / Linux 直接実行は実機未検証(コードは実装済)。Win 11 で動作確認済。"),

        ("F-011B", "アイコン自動生成・変換",
         "assets/icon.png(無ければ Pillow でダミー生成)から OSに応じた .ico (Windows 7解像度) "
         "または .icns (macOS) を自動変換。assets/ にキャッシュし mtime で再生成判定。",
         "開発者", "Phase 4", "実装済み",
         "build.py (generate_dummy_icon / png_to_ico / png_to_icns / prepare_icon)",
         "アイコンは GUI 実行時の setWindowIcon() でも反映(タイトルバー/タスクバー)。"),

        ("F-011C", "配布用ZIPパッケージ自動生成",
         "ビルド完了後 --package オプションで dist/ 配下に "
         "JPKI_Image_Signer_<OS>_<arch>_v<VER>_<mode>.zip を自動作成。"
         " --onedir なら出力フォルダ全体、--onefile なら .exe 単体を圧縮。",
         "開発者", "Phase 4", "実装済み",
         "build.py:package_distribution()",
         "実機検証で onedir版 38.7MB / onefile版 38.3MB を確認。zip展開先で独立起動成功。"),

        ("F-011D", "プライバシー警告ダイアログ (★Phase 3末で追加)",
         "署名処理開始ボタン直後に表示する強警告モーダル。.jpkiimg に氏名・住所・生年月日・性別が"
         "含まれる事実をユーザーに認識させ、チェックボックス確認とExplicit赤ボタンクリックでのみ進行可能。",
         "エンドユーザー", "Phase 3〜4", "実装済み",
         "phase3/gui/privacy_warning_dialog.py / sign_panel.py:_start_signing()",
         "Enterキー誤承認防止のためデフォルトボタンは「キャンセル」。Windowsダークモード対策のQPalette強制適用。"),

        ("F-011E", "Antivirus誤検知への対応",
         "PyInstaller生成バイナリのDefender等での誤検知に関する対処手順をREADMEに明文化。"
         " SmartScreen「詳細情報→実行」、Defender除外設定、ソースから直接実行の3経路を案内。",
         "エンドユーザー", "Phase 4", "実装済み",
         "README.md (リリース版の使い方 セクション)",
         "将来課題: コード署名証明書(Authenticode)取得で根本解決。VirusTotal事前スキャンも有効。"),

        ("F-011F", "GitHub Releases 配布対応",
         ".gitignore で個人情報含む生成物 (*.jpkiimg, sign_cert.der等) を除外。"
         "v0.1.0 タグでリリース、ZIPはRelease資産として添付する運用を確立。",
         "開発者/エンドユーザー", "Phase 4", "実装済み",
         ".gitignore / README.md (ダウンロードセクション)",
         "v0.1.0 がプロジェクト初リリース。GitHubのReleasesからZIPを取得→展開→.exe起動 で動作。"),
    ]
    write_sheet(ws1,
                "1. 機能一覧表(基本設計) — v0.1.0 (Phase 4 完了版)",
                "JPKI Image Signer 機能一覧。Phase 1〜4 全完了(緑色の行)。"
                " v0.1.0 リリース可能状態。GitHub Releases からの配布を想定。",
                headers1, rows1,
                col_widths=[10, 28, 60, 22, 12, 14, 50, 60],
                status_col_index=6)

    # =============================================================
    # Sheet 2: API仕様書(APDU + Python関数)
    # =============================================================
    ws2 = wb.create_sheet("2.API仕様書")
    headers2 = [
        "API-ID", "区分", "名称", "コマンド/シグネチャ",
        "入力", "出力/戻り値", "期待SW/例外", "用途",
        "実装ファイル", "実装関数", "実装状況",
    ]
    rows2 = [
        # ---- APDU 全部実装済 ----
        ("APDU-001", "APDU", "SELECT JPKI AP",
         "00 A4 04 0C 0A D3 92 F0 00 26 01 00 00 00 01",
         "AID 10バイト", "(なし)", "9000=成功 / 6A82=AP不在",
         "JPKI署名用アプリケーションを選択する",
         "phase2/jpki/apdu.py", "SELECT_JPKI_AP 定数", "実装済み"),

        ("APDU-002", "APDU", "SELECT 署名用証明書 EF",
         "00 A4 02 0C 02 00 01",
         "FID = 0x0001", "(なし)", "9000=成功",
         "署名用電子証明書EFを選択(PIN認証済セッション内で使用)",
         "phase2/jpki/apdu.py", "SELECT_SIGN_CERT_EF / select_ef_apdu()", "実装済み"),

        ("APDU-003", "APDU", "READ BINARY",
         "00 B0 P1 P2 Le",
         "P1P2 = オフセット, Le = 0xE0(最大読出長)",
         "EFバイナリ部分(最大Leバイト)",
         "9000=成功 / 6Cxx=Le再指定 / 6B00=範囲外(終端)",
         "EFをチャンク単位で順次読出",
         "phase2/jpki/apdu.py + phase2/jpki/session.py",
         "read_binary_apdu() / JpkiSession._read_binary_all()", "実装済み"),

        ("APDU-004", "APDU", "SELECT 署名用PIN EF",
         "00 A4 02 0C 02 00 1B",
         "FID = 0x001B", "(なし)", "9000=成功",
         "署名用PINのEFを選択",
         "phase2/jpki/apdu.py", "SELECT_SIGN_PIN_EF", "実装済み"),

        ("APDU-005", "APDU", "VERIFY (残回数確認)",
         "[Case1] 00 20 00 80 (4B)\n[Case3] 00 20 00 80 00 (5B)",
         "データなし",
         "(なし)",
         "63CX=残X回 / 6983=ロック済 / 6700=Wrong length(Case変更要)",
         "PIN試行残回数の取得(消費なし)",
         "phase2/jpki/apdu.py + session.py",
         "VERIFY_REMAINING_VARIANTS / get_pin_remaining()",
         "実装済み(2バリエーション自動フォールバック)"),

        ("APDU-006", "APDU", "VERIFY PIN",
         "00 20 00 80 Lc <PIN ASCII>",
         "PIN文字列 6〜16桁(ASCII)", "(なし)",
         "9000=成功 / 63CX=失敗(残X回) / 6983=ロック",
         "署名用PINの認証",
         "phase2/jpki/session.py", "verify_pin()",
         "実装済み(bytearray化+ゼロクリア+del)"),

        ("APDU-007", "APDU", "SELECT 署名用秘密鍵 EF",
         "00 A4 02 0C 02 00 1A",
         "FID = 0x001A", "(なし)", "9000=成功",
         "署名用秘密鍵EFを選択",
         "phase2/jpki/apdu.py", "SELECT_SIGN_KEY_EF", "実装済み"),

        ("APDU-008", "APDU", "COMPUTE DIGITAL SIGNATURE",
         "80 2A 00 80 33 <DigestInfo 51B> 00",
         "DigestInfo(SHA-256, 51B固定)",
         "RSA-2048署名(256B)",
         "9000=成功 / 6982=PIN未認証",
         "DigestInfoに対するRSA-PKCS1-v1.5署名生成",
         "phase2/jpki/session.py", "sign_digest_info()", "実装済み"),

        ("APDU-009", "APDU", "SELECT 利用者証明用証明書 EF",
         "00 A4 02 0C 02 00 0A",
         "FID = 0x000A", "(なし)", "9000=成功",
         "利用者証明用電子証明書EFを選択(PIN不要)",
         "phase2/jpki/session.py", "read_auth_certificate()", "実装済み"),

        # ---- Python関数: jpki ----
        ("FN-001", "Python関数", "JpkiSession (class)",
         "JpkiSession(reader_index=None) → ContextManager",
         "リーダー番号(オプション)",
         "セッションオブジェクト",
         "JpkiNoReaderError / JpkiCardError",
         "JPKIカード操作の高レベルクラス。with文対応で確実に切断",
         "phase2/jpki/session.py", "JpkiSession", "実装済み"),

        ("FN-002", "Python関数", "JpkiSession.assert_safe_to_attempt_pin()",
         "(threshold=3) → int|None",
         "閾値(オプション)",
         "残回数 / None=取得不能",
         "JpkiPinLockedError / JpkiPinRiskError",
         "残回数<閾値で例外。Phase 1の安全装置を例外駆動に整理",
         "phase2/jpki/session.py", "assert_safe_to_attempt_pin", "実装済み"),

        ("FN-003", "Python関数", "JpkiSession.verify_pin(pin)",
         "(pin: str) → None",
         "PIN文字列(6〜16桁ASCII)",
         "(成功時 None)",
         "JpkiPinFailedError(remaining) / JpkiPinLockedError / ValueError",
         "PIN認証。bytearray化 → APDU送信 → ゼロクリア + del",
         "phase2/jpki/session.py", "verify_pin", "実装済み"),

        ("FN-004", "Python関数", "JpkiSession.sign_digest_info(di)",
         "(bytes 51B) → bytes 256B",
         "DigestInfo (SHA-256)",
         "RSA-2048署名値",
         "JpkiPinNotVerifiedError / JpkiCardError / ValueError",
         "PIN認証済前提。COMPUTE DIGITAL SIGNATUREを実行",
         "phase2/jpki/session.py", "sign_digest_info", "実装済み"),

        ("FN-005", "Python関数", "JpkiSession.read_sign_certificate()",
         "() → bytes",
         "(なし、PIN認証済要)",
         "EF生バイト列(パディング含む)",
         "JpkiPinNotVerifiedError",
         "署名用証明書をEF生のまま返す。trim_derでパディング除去",
         "phase2/jpki/session.py", "read_sign_certificate", "実装済み"),

        ("FN-006", "Python関数", "build_digest_info_sha256(data)",
         "(bytes) → bytes (51B)",
         "任意データ",
         "DigestInfo (SHA-256)",
         "(なし)",
         "PKCS#1 v1.5 プレフィックス + SHA-256ハッシュ で51B構築",
         "phase2/jpki/digest.py", "build_digest_info_sha256", "実装済み"),

        # ---- Python関数: crypto ----
        ("FN-100", "Python関数", "build_p7s(signature, cert_der)",
         "(bytes 256B, bytes DER) → bytes",
         "JPKI署名値 + 署名者証明書",
         "PKCS#7 SignedData (detached) DER",
         "ValueError / P7sError",
         "asn1crypto.cmsで手動構築。signedAttrs無し最小構成。version='v1'",
         "phase2/crypto/p7s.py", "build_p7s", "実装済み"),

        ("FN-101", "Python関数", "verify_p7s_against_data(p7s, content)",
         "(bytes, bytes) → bool",
         "p7s + 元データ",
         "True=有効 / False=改ざん",
         "P7sVerificationError(構造異常)",
         "RSASSA-PKCS1-v1_5+SHA-256検証。改ざん検知",
         "phase2/crypto/p7s.py", "verify_p7s_against_data", "実装済み"),

        ("FN-102", "Python関数", "verify_signed_image(image, p7s)",
         "(bytes, bytes) → dict",
         "画像 + p7s",
         "{valid, signer_cn, not_valid_before/after, error}",
         "(辞書のerrorに格納)",
         "高レベルラッパー。GUI/CLIから呼ばれる入口",
         "phase2/crypto/verify.py", "verify_signed_image",
         "実装済み(SAN氏名抽出はPhase 3で追加予定)"),

        ("FN-103", "Python関数", "actual_der_length(data) / trim_der(data)",
         "(bytes) → int / bytes",
         "DERバイト列",
         "実DER長 / トリム済DER",
         "ValueError(不正DER)",
         "ASN.1 SEQUENCEヘッダから実DER長を算出。EFパディング除去に使用",
         "phase2/crypto/der_utils.py", "actual_der_length / trim_der", "実装済み"),

        ("FN-104", "Python関数", "extract_signer_cert_der(p7s)",
         "(bytes) → bytes",
         "p7s",
         "署名者証明書 DER",
         "P7sVerificationError",
         "p7sからSignerInfo.sidに対応する証明書を抽出",
         "phase2/crypto/p7s.py", "extract_signer_cert_der", "実装済み"),

        # ---- Python関数: container ----
        ("FN-200", "Python関数", "create_jpkiimg(image_path, p7s, cert_der, output_path)",
         "(Path, bytes, bytes, Path) → Path",
         "画像パス + p7s + cert + 出力先",
         ".jpkiimgパス",
         "FileNotFoundError / ValueError / OSError",
         "ZIP_STORED で .jpkiimg を作成。元画像は無加工",
         "phase2/container/writer.py", "create_jpkiimg", "実装済み"),

        ("FN-201", "Python関数", "read_jpkiimg(path)",
         "(Path) → tuple[bytes, str, bytes, bytes]",
         ".jpkiimgパス",
         "(image, image_filename, p7s, cert_der)",
         "FileNotFoundError / NotJpkiImgError / MissingEntryError",
         ".jpkiimgを読み出して4要素タプルで返す",
         "phase2/container/reader.py", "read_jpkiimg", "実装済み"),

        # ---- CLI ----
        ("CLI-001", "CLI", "phase2.cli.sign_image",
         "py -3.12 -m phase2.cli.sign_image <image> [-o <out.jpkiimg>] [--no-color]",
         "画像ファイル",
         ".jpkiimgファイル",
         "exit 0/1/2/3",
         "実カードを使った署名フロー全体を実行",
         "phase2/cli/sign_image.py", "main()", "実装済み"),

        ("CLI-002", "CLI", "phase2.cli.verify_image",
         "py -3.12 -m phase2.cli.verify_image <jpkiimg> [--no-color]",
         ".jpkiimgファイル",
         "コンソール出力",
         "exit 0=valid / 2=構造異常 / 3=改ざん検知",
         ".jpkiimgを検証して結果を表示",
         "phase2/cli/verify_image.py", "main()", "実装済み"),

        # ---- Phase 3: GUI クラス群 ----
        ("GUI-001", "PyQt6 GUI", "MainWindow",
         "MainWindow() (QMainWindow) — 800x600 / setAcceptDrops(True)",
         "(なし、エントリポイントから生成)",
         "QStackedWidget で Welcome/Sign/Verify を切替",
         "(なし、シグナルで panel と連動)",
         "ウィンドウ全体でD&Dを受付し、ファイル拡張子で自動的に sign/verify モードへ遷移",
         "phase3/gui/main_window.py", "MainWindow.dropEvent / _enter_*_mode",
         "実装済み"),

        ("GUI-002", "PyQt6 GUI", "SignPanel",
         "SignPanel() (QWidget)",
         "set_file(Path) で対象画像をセット",
         "back_requested シグナル emit",
         "(なし)",
         "署名モードのパネル: 進捗表示・PinDialog呼出・3パターンカード表示・出力フォルダを開く",
         "phase3/gui/sign_panel.py", "SignPanel", "実装済み"),

        ("GUI-003", "PyQt6 GUI", "VerifyPanel",
         "VerifyPanel() (QWidget)",
         "set_file(Path) で対象 .jpkiimg をセット",
         "back_requested シグナル emit",
         "(なし)",
         "検証モードのパネル: VerifyWorker起動 → 緑/赤/橙の3パターンカード分岐",
         "phase3/gui/verify_panel.py", "VerifyPanel", "実装済み"),

        ("GUI-004", "PyQt6 GUI", "PinDialog",
         "PinDialog(remaining: int) (QDialog)",
         "remaining=PIN残回数",
         "get_pin() -> str | None",
         "(なし)",
         "PIN入力モーダル。残回数色分け / Password EchoMode / リアルタイムバリデーション / 閉じる時クリア",
         "phase3/gui/pin_dialog.py", "PinDialog.get_pin()", "実装済み"),

        ("GUI-005", "PyQt6 GUI", "VerifyWorker (QThread)",
         "VerifyWorker(jpkiimg_path: Path) extends QThread",
         ".jpkiimgパス",
         "シグナル: result_ready(dict), error_occurred(str)",
         "QThread.run() で例外捕捉して error_occurred 発火",
         "別スレッドで read_jpkiimg + verify_signed_image を実行(GUIフリーズ防止)",
         "phase3/gui/workers.py", "VerifyWorker.run()", "実装済み"),

        ("GUI-006", "PyQt6 GUI", "SignWorker (QThread)",
         "SignWorker(image_path: Path, output_path: Path)",
         "画像パス / 出力先パス",
         "シグナル: stage_started(str), pin_needed(int), result_ready(dict), error_occurred(str, str)",
         "kind文字列で分類: pin_locked/pin_failed/pin_risk/no_reader/card_error/cancelled/unexpected",
         "JPKIカード操作を QMutex+QWaitCondition で UIスレッドと PIN同期しつつ実行",
         "phase3/gui/workers.py", "SignWorker.run() / provide_pin() / cancel()", "実装済み"),

        ("GUI-007", "PyQt6 エントリ", "phase3.app",
         "py -3.12 -m phase3.app",
         "(なし)",
         "GUIウィンドウ起動",
         "(終了コードは Qt の event ループ準拠)",
         "QApplication 生成 → APP_STYLESHEET 適用 → MainWindow.show()",
         "phase3/app.py", "main()", "実装済み"),
    ]
    write_sheet(ws2,
                "2. API仕様書(詳細設計) — v0.1.0 (Phase 4 完了版)",
                "本プロジェクトはHTTP APIを持たない。外部I/Fは「ICカードへのAPDU」、内部I/Fは「Pythonクラス/関数」で構成される。"
                " Phase 1〜2 の全要素が実機/ユニット検証済(緑色の行)。FN-102 (verify_signed_image) は SAN 氏名抽出を Phase 3 で追加予定。",
                headers2, rows2,
                col_widths=[10, 14, 38, 50, 28, 28, 28, 32, 38, 28, 18],
                status_col_index=11)

    # =============================================================
    # Sheet 3: テーブル定義書(=データ構造定義書)
    # =============================================================
    ws3 = wb.create_sheet("3.テーブル定義書")
    headers3 = [
        "構造ID", "構造種別", "格納先", "エントリ名/カラム名",
        "論理名", "型/フォーマット", "必須", "制約", "説明",
    ]
    rows3 = [
        ("CON-001", "ZIPコンテナエントリ(実装済)", ".jpkiimg (ZIP_STORED 無圧縮)",
         "target_image.<ext>", "対象画像",
         "binary (JPEG/PNG等の元ファイルそのまま)",
         "○", "一切加工しないこと(再エンコード禁止)。拡張子は元ファイルを踏襲(小文字化)、無拡張子は.bin",
         "署名対象の画像本体。SHA-256ハッシュ計算の入力。実機テストで30,634B JPEG格納成功。"),

        ("CON-002", "ZIPコンテナエントリ(実装済)", ".jpkiimg (ZIP_STORED 無圧縮)",
         "signature.p7s", "分離署名",
         "binary (PKCS#7 / CMS Detached SignedData, DERエンコード)",
         "○", "target_image のSHA-256ハッシュに対する署名であること",
         "JPKI署名用秘密鍵によるRSA-2048署名+証明書埋込。実機テストで2,250B出力を確認。signedAttrs無し最小構成。"),

        ("CON-003", "ZIPコンテナエントリ(実装済)", ".jpkiimg (ZIP_STORED 無圧縮)",
         "cert.der", "署名者証明書",
         "binary (X.509, DERエンコード, トリム済)",
         "○", "JPKI署名用電子証明書(発行元: 公的個人認証サービス)。actual_der_lengthで実DER長を切り出し済",
         "p7s内にも同cert埋込済だが、独立アクセス用に格納。実機で1,749B(EF確保3,808Bの内、実DER部分)。氏名・住所等含むため取扱注意。"),

        ("MEM-001", "メモリ構造", "DigestInfo(SHA-256)",
         "AlgorithmIdentifier.algorithm", "ハッシュアルゴリズムOID",
         "ASN.1 OBJECT IDENTIFIER",
         "○", "2.16.840.1.101.3.4.2.1 (id-sha256) 固定",
         "build_digest_info_sha256で19Bプレフィックス内に構築。"),

        ("MEM-002", "メモリ構造", "DigestInfo(SHA-256)",
         "digest", "ハッシュ値",
         "ASN.1 OCTET STRING (32バイト)",
         "○", "32バイト固定",
         "対象データのSHA-256ハッシュ値。"),

        ("MEM-003", "メモリ構造", "DigestInfo(SHA-256)",
         "(全体)", "DigestInfo総バイト数",
         "DER全体で 51バイト固定",
         "○", "プレフィックス19B + ダイジェスト32B = 51B 固定",
         "JPKI COMPUTE DIGITAL SIGNATURE に渡すペイロード。実機検証済。"),

        ("MEM-004", "メモリ構造", "X.509証明書(DER)",
         "tbsCertificate.subject.CN", "署名者識別子(★氏名ではない)",
         "ASN.1 UTF8String / PrintableString",
         "○", "JPKI仕様: ★CN は識別符号(乱数的なID)。漢字氏名はSANに格納される(MEM-006参照)",
         "GUI上の表示用に抽出する。Phase 3でSAN優先、CNフォールバックの抽出ロジックを実装予定。"),

        ("MEM-005", "メモリ構造(実機確認済)", "利用者証明用電子証明書 (EF=0x000A)",
         "Subject", "識別符号", "X.509 DistinguishedName",
         "○", "CN=英数字符号(個人を直接特定しない)",
         "実機検証で氏名・住所・生年月日・性別を含まないことを確認。実DER 1574B。"),

        ("MEM-006", "メモリ構造(実機確認済★重要)", "署名用電子証明書 (EF=0x0001)",
         "Subject", "Subject DN(★CN=識別符号)", "X.509 DistinguishedName",
         "○", "★Subject CN は『識別符号(発行日時+乱数+連番)』であり、氏名ではない。"
         " 形式例(モック・架空): 「YYYYMMDD + 連番 + ランダム文字列」(28文字程度)",
         "★旧設計書(Phase 1完了版)では「CN=氏名」と記載していたが、実機検証で誤りと判明。"
         " 漢字氏名は SubjectAltName 拡張領域に格納される(MEM-007参照)。"),

        ("MEM-007", "メモリ構造(実機確認済★最重要)", "署名用電子証明書 (EF=0x0001)",
         "extensions.subjectAltName / OtherName 群", "署名者の漢字氏名・住所・生年月日・性別・補助番号",
         "X.509v3 拡張: SubjectAltName / OtherName 形式 (JPKI 独自OID 1.2.392.200149.8.5.5.x)",
         "○",
         "★実機検証で判明: SAN は DirectoryName ではなく **OtherName** で構成される。"
         " JPKI 独自 OID マッピング:\n"
         "  - .1 氏名(漢字)         UTF8String / 6文字+全角空白(姓+名)\n"
         "  - .2 予備フィールド      6文字\n"
         "  - .3 性別                1文字 (1=男, 2=女)\n"
         "  - .4 生年月日            9文字 [元号区分][YYYYMMDD]\n"
         "  - .5 住所(漢字)         UTF8String 可変長\n"
         "  - .6 補助番号            17桁",
         "Phase 3 / Step 1 (v2) で phase2/crypto/verify.py の extract_signer_name を改修:"
         " OtherName.type_id が 1.2.392.200149.8.5.5.1 のエントリを最優先し、"
         " value(ASN.1 UTF8String) をデコードして漢字氏名として返す。"
         " 出典: 公的個人認証サービス(J-LIS) 署名用電子証明書 仕様。"),

        ("MEM-008", "メモリ構造(実機確認済)", "RSA署名値",
         "(本体)", "RSA-2048署名", "binary 256バイト",
         "○", "256B固定 (RSA-2048 / RSASSA-PKCS1-v1_5 / SHA-256)",
         "JPKIカードのCOMPUTE DIGITAL SIGNATUREの出力。実機で256Bを確認。signature_dummy.binやsignature.p7s内に格納。"),

        ("MEM-009", "メモリ構造", "PKCS#7 SignedData (detached)",
         "(全体)", "分離署名コンテンツ", "asn1crypto.cms.ContentInfo (DER)",
         "○", "version='v1', signedAttrs無し, encap_content_info.content=ABSENT",
         "実機で2,250Bを確認。PKCS#7 v1.5互換構成(asn1crypto側でversion='v1'の場合encap_content_infoはContentInfo型)。"),
    ]
    write_sheet(ws3,
                "3. テーブル定義書(詳細設計) — v0.1.0 (Phase 4 完了版)",
                "本プロジェクトはRDBMSを使用しない。データ永続化は .jpkiimg コンテナ(無圧縮ZIP)で行う。"
                " 「永続化フォーマット定義(CON-)」と「主要メモリ構造定義(MEM-)」を兼ねる。"
                " ★MEM-006/MEM-007 は Phase 2 実機検証で判明した重要事項(Subject CNは氏名ではなく識別符号、漢字氏名はSANに存在)。",
                headers3, rows3,
                col_widths=[10, 26, 36, 32, 24, 38, 8, 50, 60])

    # =============================================================
    # Sheet 4: エラー・ログ定義書
    # =============================================================
    ws4 = wb.create_sheet("4.エラー・ログ定義書")
    headers4 = [
        "エラーID", "種別", "コード/例外名", "メッセージ/出力",
        "発生条件", "ログレベル", "対処", "発生箇所",
    ]
    rows4 = [
        # APDU SW
        ("E-APDU-9000", "APDU SW", "9000",
         "(成功・出力なし)", "APDUコマンド正常終了", "INFO",
         "(対処不要)", "全APDU共通"),

        ("E-APDU-6A82", "APDU SW", "6A82",
         "署名用APが見つかりません", "SELECT AP時にAPが存在しない",
         "ERROR", "マイナンバーカードか確認", "JpkiSession.__init__"),

        ("E-APDU-6CXX", "APDU SW", "6Cxx",
         "(自動再試行・DEBUG出力)", "READ BINARY等のLeが実データ長と不一致",
         "DEBUG", "sw2を新Leとして自動再送", "_read_binary_all() / sign_digest_info()"),

        ("E-APDU-6B00", "APDU SW", "6B00",
         "(終端到達)", "READ BINARYのオフセットが範囲外",
         "DEBUG", "正常終端としてループを抜ける", "_read_binary_all()"),

        ("E-APDU-6700", "APDU SW", "6700",
         "Wrong length / Case変更フォールバック",
         "VERIFY 残回数確認 APDUのバリエーションをカードが認識しない",
         "WARN",
         "次のバリエーションに自動フォールバック。試行は消費されない。",
         "get_pin_remaining()"),

        ("E-APDU-63CX", "APDU SW", "63Cx",
         "PIN残回数: x回", "VERIFY後にPIN不一致 または 残回数照会成功",
         "WARN/INFO",
         "残回数<閾値で JpkiPinRiskError / 失敗時 JpkiPinFailedError",
         "verify_pin() / get_pin_remaining()"),

        ("E-APDU-6983", "APDU SW", "6983",
         "PINがロックされています", "残試行回数0でVERIFY実行",
         "FATAL",
         "市区町村窓口で初期化が必要 → JpkiPinLockedError",
         "verify_pin() / get_pin_remaining()"),

        ("E-APDU-6982", "APDU SW", "6982",
         "PIN未認証 / セキュリティ状態未充足",
         "VERIFY未実行のままPIN必須EF/COMPUTE呼出",
         "ERROR", "VERIFYからやり直し", "session内部"),

        # JPKI 例外
        ("E-JPKI-001", "JPKI例外", "JpkiNoReaderError",
         "ICカードリーダーが見つかりません",
         "リーダー無し / 複数台かつindex未指定 / index範囲外",
         "ERROR", "リーダー接続を確認", "JpkiSession.__init__"),

        ("E-JPKI-002", "JPKI例外", "JpkiCardError",
         "APDU/カード通信エラー",
         "SW異常 / CardConnectionException 等",
         "ERROR", "カード再挿入・SCardSvr再起動",
         "session.py 全般"),

        ("E-JPKI-003", "JPKI例外", "JpkiPinLockedError",
         "PINがロックされています",
         "SW=6983 検出 / VERIFY試行で残0",
         "FATAL", "市区町村窓口で初期化必須",
         "verify_pin() / get_pin_remaining()"),

        ("E-JPKI-004", "JPKI例外", "JpkiPinFailedError(remaining)",
         "PIN認証失敗 残N回",
         "SW=63CX after VERIFY",
         "ERROR", "PIN確認後 --check-only で残回数確認 → 慎重に再実行",
         "verify_pin()"),

        ("E-JPKI-005", "JPKI例外", "JpkiPinRiskError(remaining, threshold)",
         "残回数 X < 閾値 Y",
         "残回数 < MIN_SAFE_REMAINING(3)",
         "FATAL", "PIN確認 / カードが正しいか確認",
         "assert_safe_to_attempt_pin()"),

        ("E-JPKI-006", "JPKI例外", "JpkiPinNotVerifiedError",
         "PIN認証前にPIN必須操作が呼ばれた",
         "verify_pin成功前にsign_digest_info/read_sign_certificate呼出",
         "ERROR(プログラミングエラー)", "verify_pin成功を待つ",
         "sign_digest_info() / read_sign_certificate()"),

        # PKCS#7 例外
        ("E-P7S-001", "PKCS#7例外", "P7sError",
         "PKCS#7構造構築失敗",
         "asn1crypto側で構造組立失敗",
         "ERROR", "入力(signature/cert)を確認",
         "build_p7s()"),

        ("E-P7S-002", "PKCS#7例外", "P7sVerificationError",
         "p7s構造異常(検証以前の問題)",
         "p7sがDERでない / SignerInfoが0個または2個以上 / cert不一致 等",
         "ERROR", "p7sファイル/コンテナの破損確認",
         "verify_p7s_against_data() / extract_signer_cert_der()"),

        # コンテナ例外
        ("E-CONT-001", "コンテナ例外", "NotJpkiImgError",
         "ZIP形式でない / 開けない",
         "BadZipFile / CRCエラー",
         "ERROR", "ファイル破損確認・再取得",
         "read_jpkiimg()"),

        ("E-CONT-002", "コンテナ例外", "MissingEntryError",
         "必須エントリ欠落",
         "target_image.* / signature.p7s / cert.der のいずれかが無い",
         "ERROR", ".jpkiimg を再生成",
         "read_jpkiimg()"),

        # CLI 終了コード
        ("E-CLI-001", "プロセス終了コード", "1",
         "ファイル/リソースエラー", "ファイル不在 / pyscard未インストール 等",
         "FATAL", "メッセージに従う",
         "phase2/cli/sign_image.py / verify_image.py"),

        ("E-CLI-002", "プロセス終了コード", "2",
         "カード/コンテナ構造エラー", "JpkiCardError / NotJpkiImgError / MissingEntryError",
         "ERROR", "メッセージに従う", "phase2/cli/*"),

        ("E-CLI-003", "プロセス終了コード", "3",
         "署名検証失敗 / 安全装置作動",
         "PIN失敗 / ロック / 安全装置 / 検証FALSE",
         "ERROR/FATAL", "状況に応じて対処", "phase2/cli/*"),

        # ログ
        ("L-001", "コンソールログ", "(INFO/緑)",
         "[OK] 接続成功 / ✅ 有効な署名です",
         "正常完了時", "INFO", "(対処不要)",
         "phase2/cli/* + _terminal.py"),

        ("L-002", "コンソールログ", "(WARN/黄)",
         "[STOP] 安全装置作動 / 残回数<閾値",
         "セキュリティ関連の警告", "WARN",
         "メッセージに従う", "phase2/cli/sign_image.py"),

        ("L-003", "コンソールログ", "(ERROR/赤)",
         "❌ 検証エラー / [FATAL] PINロック",
         "致命的エラー / 改ざん検知",
         "ERROR", "メッセージに従う", "phase2/cli/*"),

        ("L-SEC-001", "セキュリティログ", "(SECURITY)",
         "(画面に何も表示せずPIN受付・APDU送信直後に bytearray ゼロクリア+del)",
         "PIN入力〜VERIFY間",
         "SECURITY", "意図的に詳細出力なし。失敗時のみSWを表示",
         "verify_pin() / sign_image.py"),

        # ユニットテスト
        ("L-TEST-001", "テストログ", "Ran 50 tests in X.XXXs / OK",
         "phase2/tests 全パス時の出力 (Phase 3で SAN OtherName抽出 5件追加)",
         "python -m unittest discover -s phase2/tests 実行時",
         "INFO", "(対処不要)",
         "phase2/tests/test_p7s.py (27件) + test_container.py (18件)"),

        # ---- Phase 3 GUI: SignWorker.error_occurred kind 分類 ----
        ("E-GUI-PIN-LOCKED", "GUI Worker error_kind", "pin_locked",
         "❌ 署名用PINがロックされています",
         "JpkiPinLockedError catch (既ロック or VERIFY試行で残0)",
         "FATAL",
         "市区町村窓口での初期化が必要(本人確認書類持参)",
         "phase3/gui/workers.py:SignWorker / phase3/gui/sign_panel.py:_show_failure_card"),

        ("E-GUI-PIN-FAILED", "GUI Worker error_kind", "pin_failed",
         "❌ PIN認証失敗 残N回",
         "JpkiPinFailedError catch",
         "ERROR",
         "PIN確認後、SignPanel の「もう一度署名する」で再実行(ただし試行は減る)",
         "phase3/gui/workers.py:SignWorker.run() VERIFY段階"),

        ("E-GUI-PIN-RISK", "GUI Worker error_kind", "pin_risk",
         "⚠ 安全装置作動 (PIN残回数<3)",
         "JpkiPinRiskError catch (assert_safe_to_attempt_pin)",
         "FATAL",
         "PIN確認 / カードが正しいか確認後に手動で再実行(--check-only相当でカウント確認推奨)",
         "phase3/gui/workers.py:SignWorker Stage 3"),

        ("E-GUI-NO-READER", "GUI Worker error_kind", "no_reader",
         "⚠ ICカードリーダーが見つかりません",
         "JpkiNoReaderError catch",
         "ERROR",
         "USB再接続 / SCardSvr サービス確認 / ドライバ確認",
         "phase3/gui/workers.py:SignWorker Stage 2"),

        ("E-GUI-CARD-ERROR", "GUI Worker error_kind", "card_error",
         "⚠ カード通信エラー",
         "JpkiCardError catch (任意ステージのSW異常)",
         "ERROR",
         "カード抜き差し / リーダー再接続 / アプリ再起動",
         "phase3/gui/workers.py:SignWorker 全ステージ"),

        ("E-GUI-CANCELLED", "GUI Worker error_kind", "cancelled",
         "(グレー文字「キャンセルされました」のみ)",
         "PinDialogでキャンセル / SignWorker.cancel() 呼出 / isInterruptionRequested=True",
         "INFO",
         "(対処不要)",
         "phase3/gui/workers.py:SignWorker / sign_panel.py"),

        ("E-GUI-UNEXPECTED", "GUI Worker error_kind", "unexpected",
         "⚠ 想定外のエラー",
         "上記いずれにも該当しない例外(ImportError / IO等)",
         "ERROR",
         "ログを確認し、開発者へ報告",
         "phase3/gui/workers.py:SignWorker run() except Exception"),

        # ---- Phase 3 GUI: コンソール無し / トーストレベル ----
        ("L-GUI-001", "GUIログ", "(stage_started)",
         "[N/8] 画像を読み込み中... 等の進捗",
         "SignWorker / VerifyWorker の各ステージ開始時",
         "INFO", "(対処不要、UI status_label で表示)",
         "phase3/gui/sign_panel.py:_on_stage_started"),

        ("L-GUI-002", "GUIログ", "(pin_needed)",
         "PinDialog 表示要求(残回数N回)",
         "SignWorker Stage 4 でUIへPIN要求",
         "INFO", "(対処不要)",
         "phase3/gui/sign_panel.py:_on_pin_needed"),
    ]
    write_sheet(ws4,
                "4. エラー・ログ定義書(詳細設計) — v0.1.0 (Phase 4 完了版)",
                "APDUステータスワード / カスタム例外階層 / プロセス終了コード / コンソールログ / セキュリティログ を一覧化。"
                " Phase 2 で 例外階層(JpkiError/PinError/CardError/ContainerError/P7sError) を整理済。",
                headers4, rows4,
                col_widths=[16, 22, 22, 38, 42, 12, 48, 38])

    # =============================================================
    # Sheet 5: アーキテクチャ図解(Mermaid)
    # =============================================================
    ws5 = wb.create_sheet("5.アーキテクチャ図解")
    headers5 = ["図ID", "図種", "タイトル", "概要", "Mermaidコード"]

    diag_overview = """graph TD
    subgraph User["ユーザー環境"]
        U[クリエイター]
        IMG[元画像 JPEG/PNG]
        OUT[.jpkiimgファイル]
    end
    subgraph App["JPKI Image Signer (Python 3.12+)"]
        GUI["GUI層 PyQt6/CTk (Phase3 未実装)"]
        CLI["CLI層 sign_image/verify_image (Phase2 完了)"]
        SIGN["署名サービス (build_p7s + create_jpkiimg) (Phase2 完了)"]
        VERIFY["検証サービス (read_jpkiimg + verify_p7s + verify_signed_image) (Phase2 完了)"]
        JPKI["JPKI通信 phase2/jpki/ (Phase2 完了)"]
        CRYPTO["暗号処理 phase2/crypto/ (Phase2 完了)"]
        ZIP["コンテナ管理 phase2/container/ (Phase2 完了)"]
    end
    subgraph HW["ハードウェア (実機検証済)"]
        PCSC["PC/SC層 Microsoft Usbccid (WUDF)"]
        READER["Alcor AU9540 系 接触型リーダー"]
        CARD["マイナンバーカード"]
    end
    U -->|D&D 予定| GUI
    U -->|現状CLI| CLI
    IMG --> CLI
    GUI --> SIGN
    GUI --> VERIFY
    CLI --> SIGN
    CLI --> VERIFY
    SIGN --> CRYPTO
    SIGN --> JPKI
    SIGN --> ZIP
    VERIFY --> CRYPTO
    VERIFY --> ZIP
    JPKI --> PCSC
    PCSC --> READER
    READER -->|接触| CARD
    ZIP --> OUT
"""

    diag_sign_seq = """sequenceDiagram
    actor U as ユーザー
    participant CLI as sign_image.py
    participant J as JpkiSession (phase2/jpki)
    participant CR as crypto/p7s
    participant Z as container/writer
    participant C as カード
    U->>CLI: py -m phase2.cli.sign_image image.jpg
    CLI->>CLI: image読込
    CLI->>J: JpkiSession() (with __enter__)
    J->>C: SELECT JPKI AP
    C-->>J: 9000
    CLI->>J: assert_safe_to_attempt_pin()
    J->>C: SELECT 署名用PIN EF
    J->>C: VERIFY 4byte (00 20 00 80)
    C-->>J: 63Cx (残x回)
    alt 残回数 < 3 or ロック
        J-->>CLI: JpkiPinRiskError / JpkiPinLockedError
        CLI-->>U: 安全装置作動・終了
    end
    CLI-->>U: 続行確認 yes/no
    U->>CLI: yes / PIN入力(getpass)
    CLI->>J: verify_pin(pin)
    J->>J: pin → bytearray
    J->>C: VERIFY PIN
    C-->>J: 9000
    J->>J: bytearray ゼロクリア + del
    CLI->>CR: build_digest_info_sha256(image)
    CR-->>CLI: DigestInfo 51B
    CLI->>J: sign_digest_info(di)
    J->>C: SELECT 署名用秘密鍵 EF
    J->>C: COMPUTE DIGITAL SIGNATURE
    C-->>J: 256B signature
    CLI->>J: read_sign_certificate()
    J->>C: SELECT 署名用cert EF + READ BINARY loop
    C-->>J: cert (3808B EF)
    CLI->>CR: trim_der(cert) → 1749B
    CLI->>CR: build_p7s(sig, cert)
    CR-->>CLI: p7s 2,250B
    CLI->>Z: create_jpkiimg(image, p7s, cert)
    Z-->>CLI: .jpkiimg 34,957B
    CLI-->>U: ✅ 完了
"""

    diag_verify_seq = """sequenceDiagram
    actor U as ユーザー
    participant CLI as verify_image.py
    participant Z as container/reader
    participant V as crypto/verify
    participant CR as crypto/p7s
    U->>CLI: py -m phase2.cli.verify_image x.jpkiimg
    CLI->>Z: read_jpkiimg(path)
    alt ZIPでない
        Z-->>CLI: NotJpkiImgError
        CLI-->>U: ❌ 不正なコンテナ
    else 必須欠落
        Z-->>CLI: MissingEntryError
        CLI-->>U: ❌ 必須エントリ欠落
    end
    Z-->>CLI: (image, name, p7s, cert)
    CLI->>V: verify_signed_image(image, p7s)
    V->>CR: verify_p7s_against_data(p7s, image)
    CR->>CR: actual_der_length / 公開鍵抽出 / RSA-PKCS1-v15 検証
    alt 検証成功
        CR-->>V: True
        V->>V: extract_signer_cert / Subject CN取得<br/>(★Phase3でSAN氏名抽出に拡張予定)
        V-->>CLI: {valid:true, signer_cn, validity}
        CLI-->>U: ✅ 有効・署名者表示
    else 改ざん検知
        CR-->>V: False
        V-->>CLI: {valid:false}
        CLI-->>U: ❌ 検証エラー
    end
"""

    diag_container = """graph LR
    subgraph JPKIIMG[".jpkiimg ファイル(無圧縮ZIP / 実装済)"]
        IMG["target_image.jpg<br/>元画像バイナリ<br/>無加工"]
        SIG["signature.p7s<br/>CMS Detached<br/>SignedData DER<br/>(2,250B 実例)"]
        CRT["cert.der<br/>X.509署名用証明書 DER<br/>(1,749B / トリム済)"]
    end
    HASH(("SHA-256ハッシュ"))
    KEY(("JPKI署名鍵<br/>(カード内)"))
    IMG -->|hash| HASH
    HASH -->|sign by| KEY
    KEY -->|生成| SIG
    CRT -.->|公開鍵で検証| SIG
"""

    diag_modules = """graph LR
    subgraph phase1["phase1/ (実機検証用CLI・完了)"]
        T1[test_01_connect.py]
        T2[test_02_read_cert.py]
        T3[test_03_sign.py]
        T4[test_04_verify_dummy.py]
    end
    subgraph phase2["phase2/ (再利用モジュール・完了)"]
        subgraph jpki["jpki/"]
            APDU[apdu.py]
            DIG[digest.py]
            SES[session.py<br/>JpkiSession]
        end
        subgraph crypto["crypto/"]
            DER[der_utils.py]
            P7S[p7s.py]
            VRF[verify.py]
        end
        subgraph container["container/"]
            WRT[writer.py]
            RDR[reader.py]
        end
        subgraph cli["cli/"]
            SI[sign_image.py]
            VI[verify_image.py]
            TRM[_terminal.py]
        end
        subgraph tests["tests/ (40テストPass)"]
            TP7S[test_p7s.py 22件]
            TCNT[test_container.py 18件]
        end
    end
    subgraph phase3["phase3/ (GUI・未実装)"]
        GUI[PyQt6/CTk]
    end
    SI --> SES
    SI --> P7S
    SI --> WRT
    VI --> RDR
    VI --> VRF
    VRF --> P7S
    P7S --> DER
    SES --> APDU
    SES --> DIG
    GUI -.-> SI
    GUI -.-> VI
"""

    diag_phase2_state = """stateDiagram-v2
    [*] --> CLI起動
    CLI起動 --> 画像読込: sign_image
    CLI起動 --> コンテナ読込: verify_image

    画像読込 --> JpkiSession接続
    JpkiSession接続 --> 残回数確認: SW=9000
    JpkiSession接続 --> [*]: NoReader/CardError

    残回数確認 --> 続行確認: 残>=3
    残回数確認 --> [*]: PinRisk/Locked (exit 3)

    続行確認 --> [*]: no (exit 0)
    続行確認 --> PIN入力: yes
    PIN入力 --> VERIFY実行
    VERIFY実行 --> [*]: PinFailed/Locked (exit 3)
    VERIFY実行 --> 署名取得: SW=9000
    署名取得 --> 証明書取得
    証明書取得 --> trim_der
    trim_der --> build_p7s
    build_p7s --> create_jpkiimg
    create_jpkiimg --> [*]: ✅ 出力完了 (exit 0)

    コンテナ読込 --> [*]: NotJpkiImg/MissingEntry (exit 2)
    コンテナ読込 --> verify_p7s: 3要素取得
    verify_p7s --> [*]: True → ✅ 表示 (exit 0)
    verify_p7s --> [*]: False → ❌ 改ざん検知 (exit 3)
"""

    # ============================================================
    # Phase 3 GUI 関連 Mermaid 図
    # ============================================================

    diag_gui_module = """graph TD
    subgraph entry["エントリ"]
        APP["phase3/app.py<br/>QApplication起動<br/>QSS適用 → MainWindow.show()"]
    end

    subgraph window["メインウィンドウ"]
        MW["MainWindow (QMainWindow)<br/>setAcceptDrops=True<br/>QStackedWidget で 3画面切替"]
        WP["WelcomePanel<br/>(D&D 待機画面)"]
        SP["SignPanel"]
        VP["VerifyPanel"]
    end

    subgraph workers["バックグラウンドワーカー"]
        SW["SignWorker (QThread)<br/>QMutex+QWaitCondition で<br/>PIN受け渡し"]
        VW["VerifyWorker (QThread)"]
    end

    subgraph dialogs["ダイアログ"]
        PD["PinDialog (QDialog modal)<br/>残回数色分け / Password入力 /<br/>QPalette でライトテーマ強制"]
    end

    subgraph backend["バックエンド (Phase 2)"]
        JPKI["phase2/jpki/<br/>JpkiSession"]
        CRYPTO["phase2/crypto/<br/>build_p7s / verify_signed_image /<br/>extract_signer_name"]
        CONT["phase2/container/<br/>create_jpkiimg / read_jpkiimg"]
    end

    APP --> MW
    MW --> WP
    MW --> SP
    MW --> VP
    SP --> SW
    VP --> VW
    SP -.PIN要求/応答.-> PD
    SW --> JPKI
    SW --> CRYPTO
    SW --> CONT
    VW --> CONT
    VW --> CRYPTO
"""

    diag_pin_sequence = """sequenceDiagram
    actor U as ユーザー
    participant SP as SignPanel<br/>(UIスレッド)
    participant SW as SignWorker<br/>(QThread)
    participant PD as PinDialog<br/>(QDialog modal)
    participant Card as JPKIカード

    U->>SP: 「署名を開始」クリック
    SP->>SW: start()
    Note right of SW: Stage 1〜3 を実行<br/>(画像読込/接続/残回数確認)
    SW->>SP: stage_started("[1/8] 画像を読込中...")
    SW->>SP: stage_started("[2/8] JPKIカードに接続中...")
    SW->>SP: stage_started("[3/8] PIN残回数を確認中...")
    SW->>Card: assert_safe_to_attempt_pin()
    Card-->>SW: 残N回 (N >= 3)

    SW->>SP: pin_needed.emit(N)
    Note over SW: ★ _pin_condition.wait() で待機

    SP->>PD: PinDialog(N).exec()
    PD->>U: 残回数バッジ表示<br/>PIN入力受付
    U->>PD: PIN入力 + 認証ボタン
    PD-->>SP: get_pin() returns "PIN文字列"
    SP->>SW: provide_pin(pin_str)
    Note right of SW: bytearray化 → wakeAll()<br/>→ 待機解除

    SW->>SW: bytes(pin_bytes).decode() → 一時str
    SW->>Card: VERIFY PIN
    Card-->>SW: 9000
    SW->>SW: bytearray ゼロクリア + del

    SW->>SP: stage_started("[5/8] PIN認証中...")
    SW->>Card: COMPUTE DIGITAL SIGNATURE
    Card-->>SW: 256B 署名
    SW->>Card: READ BINARY (cert)
    Card-->>SW: cert.der

    Note over SW: PKCS#7構築 + .jpkiimg作成

    SW->>SP: result_ready({output_path, signer_name, ...})
    SP->>U: ✅ 緑カード表示<br/>「📂 出力先フォルダを開く」ボタン
"""

    diag_dnd_dispatch = """flowchart TD
    Start([ユーザーがファイルD&D])
    Drag{ホバー中}
    DropEvt[dropEvent 発火]
    Single{単一ファイル?}
    Ext{拡張子判別}
    Sign[SignPanel.set_file<br/>+ stack 切替]
    Verify[VerifyPanel.set_file<br/>+ stack 切替]
    Reject[QMessageBox.warning<br/>「非対応のファイル形式」]
    Status[StatusBar 5秒間表示]

    Start --> Drag
    Drag -->|対応形式| Highlight["背景=DBEAFE<br/>枠=点線青"]
    Drag -->|非対応| Cursor[カーソル「禁止」]
    Highlight --> DropEvt
    DropEvt --> Single
    Single -->|No| Reject
    Single -->|Yes| Ext
    Ext -->|.jpg/.jpeg/.png| Sign
    Ext -->|.jpkiimg| Verify
    Ext -->|その他| Reject
    Reject --> Status
"""

    rows5 = [
        ("DIAG-001", "graph TD", "システム全体構成図(Phase 3完了版)",
         "Phase 1〜3 全レイヤを反映。Phase 4 (.exe化) のみ残。",
         diag_overview),

        ("DIAG-002", "sequenceDiagram", "署名モード シーケンス図(実装ベース)",
         "phase2/cli/sign_image.py の実フロー。実機テストで通った全ステップとサイズ実例(30634/51/256/3808/1749/2250/34957)を反映。",
         diag_sign_seq),

        ("DIAG-003", "sequenceDiagram", "検証モード シーケンス図",
         "phase2/cli/verify_image.py の実フロー。改ざん検知分岐を含む。Phase 3でSAN氏名抽出を追加予定。",
         diag_verify_seq),

        ("DIAG-004", "graph LR", ".jpkiimg コンテナ構造図",
         "コンテナの3エントリと暗号的依存関係(画像→ハッシュ→署名→証明書)を実数値付きで表示。",
         diag_container),

        ("DIAG-005", "graph LR", "モジュール構成図(Phase 2完了)",
         "phase1/phase2/phase3 のディレクトリ構成と依存関係。tests は40件Pass。",
         diag_modules),

        ("DIAG-006", "stateDiagram-v2", "Phase 2 全体状態遷移図",
         "CLI 2本(sign_image, verify_image)の全分岐と終了コード。実機検証で全経路の動作を確認済。",
         diag_phase2_state),

        # ---- Phase 3 専用図 ----
        ("DIAG-007", "graph TD", "Phase 3 GUI モジュール構成図",
         "phase3/app.py エントリから 各 Panel / Worker / Dialog を経由して Phase 2 バックエンドを呼ぶ全体構造。",
         diag_gui_module),

        ("DIAG-008", "sequenceDiagram", "PIN同期(QMutex/QWaitCondition)シーケンス図 ★Phase 3 核心",
         "SignWorker (QThread) が pin_needed を emit して UIスレッドからのPIN到着を _pin_condition.wait() で待機。"
         " SignPanel が PinDialog を開き、ユーザー入力を provide_pin() でワーカーに渡し wakeAll() で待機解除。"
         " ワーカー側で bytearray化 → APDU送信 → ゼロクリアの安全フロー。",
         diag_pin_sequence),

        ("DIAG-009", "flowchart TD", "D&D による自動モード判別フローチャート",
         "MainWindow が単一ファイル D&D を受け取り、拡張子で sign/verify を自動判別。"
         " 非対応ファイルは警告ダイアログで拒否。視覚フィードバック(背景色変化)も含む。",
         diag_dnd_dispatch),
    ]
    write_sheet(ws5,
                "5. アーキテクチャ図解(Mermaid記法) — v0.1.0 (Phase 4 完了版)",
                "Mermaid記法のテキストとして保持。https://mermaid.live/ で即時レンダリング可能。"
                " Phase 2完了時点の実装(40テストPass + 実機E2E成功)を全反映。",
                headers5, rows5,
                col_widths=[10, 20, 36, 60, 90])

    for r in range(5, 5 + len(rows5)):
        ws5.row_dimensions[r].height = 340

    # =============================================================
    wb.save(OUTPUT_PATH)
    print(f"[OK] 設計書を生成しました: {OUTPUT_PATH}")
    print(f"     シート数: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"       - {name}  (rows={ws.max_row}, cols={ws.max_column})")


if __name__ == "__main__":
    build_workbook()
